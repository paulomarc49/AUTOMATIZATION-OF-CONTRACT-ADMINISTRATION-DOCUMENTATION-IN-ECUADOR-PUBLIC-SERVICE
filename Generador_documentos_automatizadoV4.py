import re
import fitz  # PyMuPDF
from jinja2 import Template
from datetime import datetime, timedelta
import locale
import openpyxl

# === Paso 1: Extraer texto del PDF ===
def extraer_texto_pdf(ruta_pdf):
    doc = fitz.open(ruta_pdf)
    texto_total = ""
    for pagina in doc:
        texto_total += pagina.get_text("text") + "\n"
    doc.close()
    return texto_total

# === Paso 2: Extraer campos relevantes del texto ===
def extraer_campos(texto):
    datos = {}

    # Número de orden
    orden_match = re.search(r"IC-UTA\.DAD-\d{3}-\d{4}", texto)
    datos["numero_orden"] = orden_match.group(0) if orden_match else "NO ENCONTRADO"

    # Objeto de contratación
    objeto_match = re.search(
        r"OBJETO\s+DE\s+CONTRATACIÓN:.*?proveer del:\s*(.*?)\s*,\s*conforme el siguiente detalle:",
        texto, re.DOTALL | re.IGNORECASE
    )
    if objeto_match:
        datos["objeto_contratacion"] = objeto_match.group(1).replace("\n", " ").strip()
    else:
        datos["objeto_contratacion"] = "NO ENCONTRADO"

    # Proveedor
    proveedor_match = re.search(r"PROVEEDOR:\s*(.+)", texto)
    datos["proveedor"] = proveedor_match.group(1).strip() if proveedor_match else "NO ENCONTRADO"

    # Plazo de ejecución (dos patrones posibles)
    plazo_match = re.search(
        r"plazo\s+para\s+la\s+prestación.*?es\s+de\s+(\d+)\s*d[ií]as",
        texto, re.IGNORECASE | re.DOTALL
    )
    if not plazo_match:
        plazo_match = re.search(
            r"PLAZO\s+DE\s+EJECUCIÓN:.*?es\s*de\s*(\d+)\s*d[ií]as",
            texto, re.IGNORECASE | re.DOTALL
        )

    if plazo_match:
        datos["plazo_dias"] = int(plazo_match.group(1))
        print(f"📌 Plazo detectado: {datos['plazo_dias']} días")
    else:
        datos["plazo_dias"] = 15
        print("⚠️ No se detectó el plazo, se usará valor por defecto: 15 días")

    # Establecer el locale en español
    try:
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')  # Linux / macOS
    except:
        try:
            locale.setlocale(locale.LC_TIME, 'es_ES')     # Windows (algunas versiones)
        except:
            try:
                locale.setlocale(locale.LC_TIME, 'Spanish_Spain')  # Otra alternativa en Windows
            except:
                print("⚠️ No se pudo establecer el locale en español. Los meses podrían aparecer en inglés.")


    # Fecha de suscripción (manual)
    fecha_input = input("🔹 Ingresa la fecha de suscripción (formato DD/MM/AAAA): ")
    try:
        fecha_suscripcion = datetime.strptime(fecha_input.strip(), "%d/%m/%Y")
        datos["fecha_suscripcion"] = fecha_suscripcion.strftime("%d de %B de %Y")
        fecha_limite = fecha_suscripcion + timedelta(days=datos["plazo_dias"])
        datos["fecha_limite"] = fecha_limite.strftime("%d de %B de %Y")
    except ValueError:
        datos["fecha_suscripcion"] = "NO INGRESADA"
        datos["fecha_limite"] = "NO CALCULADA"

        # === Valor sin IVA flexible basado en bloque UNIDAD seguido de múltiples montos ($) ===
    unidad_bloque_match = re.search(
        r"UNIDAD\s+\d+(?:\s+\$[0-9\.,]+){3,10}",  # UNIDAD + número + 3 a 10 valores $
        texto, re.IGNORECASE
    )

    if unidad_bloque_match:
        bloque = unidad_bloque_match.group(0)

        # Imprimir para revisión si lo deseas
        # print(f"\n🧾 Bloque detectado:\n{bloque}\n")

        montos = re.findall(r"\$[0-9\.,]+", bloque)

        if montos:
            ultimo_valor = montos[-1].replace(".", "").replace(",", ".").replace("$", "").strip()
            try:
                datos["valor_sin_iva"] = float(ultimo_valor)
            except ValueError:
                datos["valor_sin_iva"] = "FORMATO INVÁLIDO"
        else:
            datos["valor_sin_iva"] = "NO DETECTADO"


    # === Fecha de la orden de compra (detectada automáticamente) ===
    fecha_orden_match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", texto)
    if fecha_orden_match:
        try:
            fecha_orden = datetime.strptime(fecha_orden_match.group(1), "%d/%m/%Y")
            datos["fecha_orden_compra"] = fecha_orden.strftime("%d de %B de %Y")
        except:
            datos["fecha_orden_compra"] = "FORMATO INVÁLIDO"
    else:
        datos["fecha_orden_compra"] = "NO DETECTADA"

    # === Número de Certificación Presupuestaria ===
    cert_match = re.search(
        r"N[ÚU]MERO\s+DE\s+CERTIFICACI[ÓO]N\s+PRESUPUESTARIA\s*:\s*(\d+)", texto, re.IGNORECASE
    )
    if cert_match:
        datos["certificacion_presupuestaria"] = cert_match.group(1)
    else:
        datos["certificacion_presupuestaria"] = "NO DETECTADA"

    # Extras

    datos["fecha_actual"] = datetime.today().strftime("%d de %B de %Y")
    mas_un_dia = fecha_suscripcion + timedelta(days=1)
    datos["fecha_expediente_admin_bienes"] = mas_un_dia.strftime("%d de %B de %Y")
    datos["firmante_admin"] = "Mg. Luis Marcial"
    datos["correo_admin"] = "lp.marcial@uta.edu.ec"

    return datos

# === Paso 3: Mostrar los datos en consola ===
def imprimir_datos(datos):
    print("\n📋 DATOS DETECTADOS:")
    for clave, valor in datos.items():
        print(f"{clave}: {valor}")


def agregar_a_excel(datos, ruta_excel):
    wb = openpyxl.load_workbook(ruta_excel)
    ws = wb.active

    nueva_fila = [
        datos.get("numero_orden", ""),
        datos.get("proveedor", ""),
        datos.get("objeto_contratacion", ""),
        datos.get("plazo_dias", ""),
        datos.get("fecha_suscripcion", ""),
        datos.get("fecha_limite", ""),
        datos.get("fecha_orden_compra", ""),
        datos.get("certificacion_presupuestaria", ""),
        datos.get("valor_sin_iva", ""),
    ]

    # Sobrescribir 
    ws.cell(row=1,  column=2,  value=datos.get("numero_orden", "")                  )
    ws.cell(row=2,  column=2,  value=datos.get("objeto_contratacion", "")           )
    ws.cell(row=3,  column=2,  value=datos.get("proveedor", "")                     )
    ws.cell(row=4,  column=2,  value=datos.get("plazo_dias", "")                    )
    ws.cell(row=5,  column=2,  value=datos.get("valor_sin_iva", "")                 )
    ws.cell(row=9,  column=2,  value=datos.get("fecha_suscripcion", "")             )
    ws.cell(row=15, column=2,  value=datos.get("fecha_limite", "")                  )
    ws.cell(row=16, column=2,  value=datos.get("certificacion_presupuestaria", "")  )
    ws.cell(row=8,  column=2,  value=datos.get("fecha_orden_compra", "")            )
    ws.cell(row=14, column=2,  value=datos["fecha_actual"]                          )
    ws.cell(row=10, column=2,  value=datos["fecha_expediente_admin_bienes"]         )
    ws.cell(row=11, column=2,  value=datos["fecha_expediente_admin_bienes"]         )

    wb.save(ruta_excel)
    print(f"✅ Datos escritos en la fila en {ruta_excel}")

# === Main ===
if __name__ == "__main__":
    ruta_pdf = "C:/Users/Luis/Desktop/Administración de orden de compra/MANTENIMIENTO_AIRES_ACONDICIONADOS_BIBLIOTECA/CHECK_LIST/1) Orden de compra IC-UTA.DAD-056-2025.pdf"
    #ruta_pdf = "C:/Users/Luis/Downloads/1)ic-uta.dad-055-2025_mant._gas_liquuado-signed-signed_(1)-signed.pdf"
    #ruta_pdf = "C:/Users/Luis/Downloads/EjemploCarla.pdf"
    ruta_excel = "C:/Users/Luis/Desktop/Transformación Digital/DB_Administracion_Ordenes_de_Compra_V1.xlsx"
    
    texto = extraer_texto_pdf(ruta_pdf)
    datos = extraer_campos(texto)
    imprimir_datos(datos)
    agregar_a_excel(datos, ruta_excel)





